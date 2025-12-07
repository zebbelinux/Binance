"""
Rapor Üretici
Excel ve PDF raporları oluşturma
"""

from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
import os
import json
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

@dataclass
class ReportConfig:
    """Rapor konfigürasyonu"""
    output_dir: str = "reports"
    include_charts: bool = True
    include_trades: bool = True
    include_performance: bool = True
    include_ai_analysis: bool = True
    chart_style: str = "default"  # default, dark, light

class ReportGenerator:
    """Rapor üretici sınıfı"""
    
    def __init__(self, config: ReportConfig = None):
        self.logger = logging.getLogger(__name__)
        self.config = config or ReportConfig()
        
        # Çıktı klasörünü oluştur
        os.makedirs(self.config.output_dir, exist_ok=True)
        
        self.logger.info("Rapor üretici başlatıldı")
    
    def generate_excel_report(self, 
                            backtest_results: Dict[str, Any],
                            performance_analysis: Dict[str, Any],
                            monte_carlo_results: Dict[str, Any] = None,
                            filename: str = None) -> str:
        """Excel raporu oluştur"""
        try:
            if not OPENPYXL_AVAILABLE:
                self.logger.error("openpyxl paketi bulunamadı")
                return None
            
            if not filename:
                filename = f"trading_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            
            # Varsayılan sheet'i kaldır
            wb.remove(wb.active)
            
            # 1. Özet sayfası
            self._create_summary_sheet(wb, backtest_results, performance_analysis)
            
            # 2. Performans metrikleri
            self._create_performance_sheet(wb, performance_analysis)
            
            # 3. Trade detayları
            if self.config.include_trades:
                self._create_trades_sheet(wb, backtest_results)
            
            # 4. Equity curve
            self._create_equity_curve_sheet(wb, backtest_results)
            
            # 5. Monte Carlo analizi
            if monte_carlo_results and self.config.include_charts:
                self._create_monte_carlo_sheet(wb, monte_carlo_results)
            
            # 6. Risk analizi
            self._create_risk_analysis_sheet(wb, performance_analysis)
            
            # 7. Zaman bazlı analiz
            self._create_time_analysis_sheet(wb, performance_analysis)
            
            # Excel dosyasını kaydet
            wb.save(filepath)
            
            self.logger.info(f"Excel raporu oluşturuldu: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Excel raporu oluşturma hatası: {e}")
            return None
    
    def generate_pdf_report(self, 
                          backtest_results: Dict[str, Any],
                          performance_analysis: Dict[str, Any],
                          monte_carlo_results: Dict[str, Any] = None,
                          filename: str = None) -> str:
        """PDF raporu oluştur"""
        try:
            if not REPORTLAB_AVAILABLE:
                self.logger.error("reportlab paketi bulunamadı")
                return None
            
            if not filename:
                filename = f"trading_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # PDF dokümanı oluştur
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph("Trading Bot Performans Raporu", title_style))
            story.append(Spacer(1, 20))
            
            # Özet bilgiler
            self._add_summary_section(story, backtest_results, performance_analysis)
            
            # Performans metrikleri
            self._add_performance_section(story, performance_analysis)
            
            # Risk analizi
            self._add_risk_section(story, performance_analysis)
            
            # Trade analizi
            if self.config.include_trades:
                self._add_trades_section(story, backtest_results)
            
            # Monte Carlo analizi
            if monte_carlo_results:
                self._add_monte_carlo_section(story, monte_carlo_results)
            
            # PDF'i oluştur
            doc.build(story)
            
            self.logger.info(f"PDF raporu oluşturuldu: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"PDF raporu oluşturma hatası: {e}")
            return None
    
    def _create_summary_sheet(self, wb, backtest_results: Dict[str, Any], 
                            performance_analysis: Dict[str, Any]):
        """Özet sayfası oluştur"""
        try:
            ws = wb.create_sheet("Özet", 0)
            
            # Başlık
            ws['A1'] = "Trading Bot Performans Raporu"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:D1')
            
            # Temel metrikler
            basic_metrics = performance_analysis.get('basic_metrics', {})
            
            summary_data = [
                ["Metrik", "Değer"],
                ["Toplam Getiri", f"{basic_metrics.get('total_return', 0):.2%}"],
                ["Yıllık Getiri", f"{basic_metrics.get('annualized_return', 0):.2%}"],
                ["Volatilite", f"{basic_metrics.get('volatility', 0):.2%}"],
                ["Sharpe Ratio", f"{basic_metrics.get('sharpe_ratio', 0):.2f}"],
                ["Max Drawdown", f"{basic_metrics.get('max_drawdown', 0):.2%}"],
                ["Win Rate", f"{basic_metrics.get('win_rate', 0):.1f}%"],
                ["Profit Factor", f"{basic_metrics.get('profit_factor', 0):.2f}"],
                ["Toplam Trade", f"{basic_metrics.get('total_trades', 0)}"],
                ["Kazanan Trade", f"{basic_metrics.get('winning_trades', 0)}"],
                ["Kaybeden Trade", f"{basic_metrics.get('losing_trades', 0)}"]
            ]
            
            # Veriyi sayfaya ekle
            for row in summary_data:
                ws.append(row)
            
            # Stil uygula
            self._apply_table_style(ws, 1, len(summary_data), 1, 2)
            
        except Exception as e:
            self.logger.error(f"Özet sayfası oluşturma hatası: {e}")
    
    def _create_performance_sheet(self, wb, performance_analysis: Dict[str, Any]):
        """Performans sayfası oluştur"""
        try:
            ws = wb.create_sheet("Performans Metrikleri")
            
            # Risk metrikleri
            risk_metrics = performance_analysis.get('risk_metrics', {})
            
            risk_data = [
                ["Risk Metriği", "Değer"],
                ["VaR (95%)", f"{risk_metrics.get('var_95', 0):.2%}"],
                ["VaR (99%)", f"{risk_metrics.get('var_99', 0):.2%}"],
                ["CVaR (95%)", f"{risk_metrics.get('cvar_95', 0):.2%}"],
                ["CVaR (99%)", f"{risk_metrics.get('cvar_99', 0):.2%}"],
                ["Expected Shortfall (95%)", f"{risk_metrics.get('expected_shortfall_95', 0):.2%}"],
                ["Expected Shortfall (99%)", f"{risk_metrics.get('expected_shortfall_99', 0):.2%}"],
                ["Skewness", f"{risk_metrics.get('skewness', 0):.2f}"],
                ["Kurtosis", f"{risk_metrics.get('kurtosis', 0):.2f}"]
            ]
            
            for row in risk_data:
                ws.append(row)
            
            self._apply_table_style(ws, 1, len(risk_data), 1, 2)
            
        except Exception as e:
            self.logger.error(f"Performans sayfası oluşturma hatası: {e}")
    
    def _create_trades_sheet(self, wb, backtest_results: Dict[str, Any]):
        """Trade detayları sayfası oluştur"""
        try:
            ws = wb.create_sheet("Trade Detayları")
            
            trades = backtest_results.get('results', {}).get('trades', [])
            if not trades:
                ws['A1'] = "Trade verisi bulunamadı"
                return
            
            # Trade verilerini DataFrame'e çevir
            df = pd.DataFrame(trades)
            
            # Sütun başlıkları
            headers = ['ID', 'Sembol', 'Yön', 'Büyüklük', 'Giriş Fiyatı', 'Çıkış Fiyatı',
                      'Giriş Zamanı', 'Çıkış Zamanı', 'P&L', 'Komisyon', 'Slippage',
                      'Net P&L', 'Kapanış Sebebi', 'Strateji']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Veri satırları
            for row_idx, trade in enumerate(trades, 2):
                ws.cell(row=row_idx, column=1, value=trade.get('id', ''))
                ws.cell(row=row_idx, column=2, value=trade.get('symbol', ''))
                ws.cell(row=row_idx, column=3, value=trade.get('side', ''))
                ws.cell(row=row_idx, column=4, value=trade.get('size', 0))
                ws.cell(row=row_idx, column=5, value=trade.get('entry_price', 0))
                ws.cell(row=row_idx, column=6, value=trade.get('exit_price', 0))
                ws.cell(row=row_idx, column=7, value=str(trade.get('entry_time', '')))
                ws.cell(row=row_idx, column=8, value=str(trade.get('exit_time', '')))
                ws.cell(row=row_idx, column=9, value=trade.get('realized_pnl', 0))
                ws.cell(row=row_idx, column=10, value=trade.get('commission', 0))
                ws.cell(row=row_idx, column=11, value=trade.get('slippage', 0))
                ws.cell(row=row_idx, column=12, value=trade.get('net_pnl', 0))
                ws.cell(row=row_idx, column=13, value=trade.get('close_reason', ''))
                ws.cell(row=row_idx, column=14, value=trade.get('strategy_name', ''))
            
            # Stil uygula
            self._apply_table_style(ws, 1, len(trades) + 1, 1, len(headers))
            
        except Exception as e:
            self.logger.error(f"Trade sayfası oluşturma hatası: {e}")
    
    def _create_equity_curve_sheet(self, wb, backtest_results: Dict[str, Any]):
        """Equity curve sayfası oluştur"""
        try:
            ws = wb.create_sheet("Equity Curve")
            
            equity_curve = backtest_results.get('results', {}).get('equity_curve', [])
            if not equity_curve:
                ws['A1'] = "Equity curve verisi bulunamadı"
                return
            
            # Veri hazırla
            timestamps = [e['timestamp'] for e in equity_curve]
            equity_values = [e['equity'] for e in equity_curve]
            balances = [e['balance'] for e in equity_curve]
            unrealized_pnls = [e['unrealized_pnl'] for e in equity_curve]
            
            # Başlıklar
            ws['A1'] = "Zaman"
            ws['B1'] = "Equity"
            ws['C1'] = "Balance"
            ws['D1'] = "Unrealized P&L"
            
            # Veri satırları
            for i, (timestamp, equity, balance, unrealized) in enumerate(zip(timestamps, equity_values, balances, unrealized_pnls), 2):
                ws.cell(row=i, column=1, value=timestamp)
                ws.cell(row=i, column=2, value=equity)
                ws.cell(row=i, column=3, value=balance)
                ws.cell(row=i, column=4, value=unrealized)
            
            # Grafik oluştur
            if self.config.include_charts:
                chart = LineChart()
                chart.title = "Equity Curve"
                chart.style = 13
                chart.y_axis.title = 'Equity'
                chart.x_axis.title = 'Time'
                
                data = Reference(ws, min_col=2, min_row=1, max_row=len(equity_curve) + 1)
                chart.add_data(data, titles_from_data=True)
                
                ws.add_chart(chart, "F2")
            
            # Stil uygula
            self._apply_table_style(ws, 1, len(equity_curve) + 1, 1, 4)
            
        except Exception as e:
            self.logger.error(f"Equity curve sayfası oluşturma hatası: {e}")
    
    def _create_monte_carlo_sheet(self, wb, monte_carlo_results: Dict[str, Any]):
        """Monte Carlo analizi sayfası oluştur"""
        try:
            ws = wb.create_sheet("Monte Carlo Analizi")
            
            statistical_analysis = monte_carlo_results.get('statistical_analysis', {})
            
            # İstatistiksel analiz verileri
            stats_data = [
                ["Metrik", "Ortalama", "Std Sapma", "Min", "Max", "P5", "P25", "P50", "P75", "P95"]
            ]
            
            for metric, stats in statistical_analysis.items():
                if isinstance(stats, dict) and 'mean' in stats:
                    row = [
                        metric,
                        f"{stats.get('mean', 0):.4f}",
                        f"{stats.get('std', 0):.4f}",
                        f"{stats.get('min', 0):.4f}",
                        f"{stats.get('max', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p5', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p25', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p50', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p75', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p95', 0):.4f}"
                    ]
                    stats_data.append(row)
            
            for row in stats_data:
                ws.append(row)
            
            self._apply_table_style(ws, 1, len(stats_data), 1, 10)
            
        except Exception as e:
            self.logger.error(f"Monte Carlo sayfası oluşturma hatası: {e}")
    
    def _create_risk_analysis_sheet(self, wb, performance_analysis: Dict[str, Any]):
        """Risk analizi sayfası oluştur"""
        try:
            ws = wb.create_sheet("Risk Analizi")
            
            drawdown_analysis = performance_analysis.get('drawdown_analysis', {})
            drawdown_stats = drawdown_analysis.get('drawdown_stats', {})
            
            risk_data = [
                ["Risk Metriği", "Değer"],
                ["Max Drawdown", f"{drawdown_stats.get('max_drawdown', 0):.2%}"],
                ["Ortalama Drawdown", f"{drawdown_stats.get('avg_drawdown', 0):.2%}"],
                ["Drawdown Std Sapma", f"{drawdown_stats.get('drawdown_std', 0):.2%}"],
                ["Drawdown Sayısı", f"{drawdown_stats.get('num_drawdowns', 0)}"],
                ["Max Drawdown Süresi", f"{drawdown_stats.get('max_drawdown_duration', 0)} gün"]
            ]
            
            for row in risk_data:
                ws.append(row)
            
            self._apply_table_style(ws, 1, len(risk_data), 1, 2)
            
        except Exception as e:
            self.logger.error(f"Risk analizi sayfası oluşturma hatası: {e}")
    
    def _create_time_analysis_sheet(self, wb, performance_analysis: Dict[str, Any]):
        """Zaman bazlı analiz sayfası oluştur"""
        try:
            ws = wb.create_sheet("Zaman Bazlı Analiz")
            
            time_analysis = performance_analysis.get('time_analysis', {})
            
            # Aylık performans
            monthly_returns = time_analysis.get('monthly_returns', {})
            monthly_data = [
                ["Aylık Performans", ""],
                ["Ortalama Getiri", f"{monthly_returns.get('mean', 0):.2%}"],
                ["Std Sapma", f"{monthly_returns.get('std', 0):.2%}"],
                ["Min Getiri", f"{monthly_returns.get('min', 0):.2%}"],
                ["Max Getiri", f"{monthly_returns.get('max', 0):.2%}"],
                ["Pozitif Aylar", f"{monthly_returns.get('positive_months', 0)}"],
                ["Negatif Aylar", f"{monthly_returns.get('negative_months', 0)}"]
            ]
            
            for row in monthly_data:
                ws.append(row)
            
            # Haftalık performans
            weekly_returns = time_analysis.get('weekly_returns', {})
            weekly_data = [
                ["", ""],
                ["Haftalık Performans", ""],
                ["Ortalama Getiri", f"{weekly_returns.get('mean', 0):.2%}"],
                ["Std Sapma", f"{weekly_returns.get('std', 0):.2%}"],
                ["Min Getiri", f"{weekly_returns.get('min', 0):.2%}"],
                ["Max Getiri", f"{weekly_returns.get('max', 0):.2%}"],
                ["Pozitif Haftalar", f"{weekly_returns.get('positive_weeks', 0)}"],
                ["Negatif Haftalar", f"{weekly_returns.get('negative_weeks', 0)}"]
            ]
            
            for row in weekly_data:
                ws.append(row)
            
            self._apply_table_style(ws, 1, len(monthly_data) + len(weekly_data), 1, 2)
            
        except Exception as e:
            self.logger.error(f"Zaman analizi sayfası oluşturma hatası: {e}")
    
    def _apply_table_style(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """Tablo stili uygula"""
        try:
            # Başlık stili
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=start_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Kenarlık
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Sütun genişliği
            for col in range(start_col, end_col + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
        except Exception as e:
            self.logger.error(f"Tablo stili uygulama hatası: {e}")
    
    def _add_summary_section(self, story, backtest_results: Dict[str, Any], 
                           performance_analysis: Dict[str, Any]):
        """PDF özet bölümü ekle"""
        try:
            basic_metrics = performance_analysis.get('basic_metrics', {})
            
            summary_data = [
                ["Metrik", "Değer"],
                ["Toplam Getiri", f"{basic_metrics.get('total_return', 0):.2%}"],
                ["Yıllık Getiri", f"{basic_metrics.get('annualized_return', 0):.2%}"],
                ["Volatilite", f"{basic_metrics.get('volatility', 0):.2%}"],
                ["Sharpe Ratio", f"{basic_metrics.get('sharpe_ratio', 0):.2f}"],
                ["Max Drawdown", f"{basic_metrics.get('max_drawdown', 0):.2%}"],
                ["Win Rate", f"{basic_metrics.get('win_rate', 0):.1f}%"],
                ["Profit Factor", f"{basic_metrics.get('profit_factor', 0):.2f}"],
                ["Toplam Trade", f"{basic_metrics.get('total_trades', 0)}"]
            ]
            
            table = Table(summary_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Performans Özeti", getSampleStyleSheet()['Heading2']))
            story.append(table)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            self.logger.error(f"PDF özet bölümü ekleme hatası: {e}")
    
    def _add_performance_section(self, story, performance_analysis: Dict[str, Any]):
        """PDF performans bölümü ekle"""
        try:
            risk_metrics = performance_analysis.get('risk_metrics', {})
            
            risk_data = [
                ["Risk Metriği", "Değer"],
                ["VaR (95%)", f"{risk_metrics.get('var_95', 0):.2%}"],
                ["VaR (99%)", f"{risk_metrics.get('var_99', 0):.2%}"],
                ["CVaR (95%)", f"{risk_metrics.get('cvar_95', 0):.2%}"],
                ["CVaR (99%)", f"{risk_metrics.get('cvar_99', 0):.2%}"],
                ["Skewness", f"{risk_metrics.get('skewness', 0):.2f}"],
                ["Kurtosis", f"{risk_metrics.get('kurtosis', 0):.2f}"]
            ]
            
            table = Table(risk_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Risk Metrikleri", getSampleStyleSheet()['Heading2']))
            story.append(table)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            self.logger.error(f"PDF performans bölümü ekleme hatası: {e}")
    
    def _add_risk_section(self, story, performance_analysis: Dict[str, Any]):
        """PDF risk bölümü ekle"""
        try:
            drawdown_analysis = performance_analysis.get('drawdown_analysis', {})
            drawdown_stats = drawdown_analysis.get('drawdown_stats', {})
            
            risk_data = [
                ["Drawdown Metriği", "Değer"],
                ["Max Drawdown", f"{drawdown_stats.get('max_drawdown', 0):.2%}"],
                ["Ortalama Drawdown", f"{drawdown_stats.get('avg_drawdown', 0):.2%}"],
                ["Drawdown Sayısı", f"{drawdown_stats.get('num_drawdowns', 0)}"],
                ["Max Drawdown Süresi", f"{drawdown_stats.get('max_drawdown_duration', 0)} gün"]
            ]
            
            table = Table(risk_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Drawdown Analizi", getSampleStyleSheet()['Heading2']))
            story.append(table)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            self.logger.error(f"PDF risk bölümü ekleme hatası: {e}")
    
    def _add_trades_section(self, story, backtest_results: Dict[str, Any]):
        """PDF trade bölümü ekle"""
        try:
            trades = backtest_results.get('results', {}).get('trades', [])
            if not trades:
                story.append(Paragraph("Trade verisi bulunamadı", getSampleStyleSheet()['Normal']))
                return
            
            # İlk 10 trade'i göster
            recent_trades = trades[:10]
            
            trade_data = [["ID", "Sembol", "Yön", "P&L", "Sebep"]]
            
            for trade in recent_trades:
                trade_data.append([
                    trade.get('id', '')[:8] + '...',
                    trade.get('symbol', ''),
                    trade.get('side', ''),
                    f"{trade.get('net_pnl', 0):.2f}",
                    trade.get('close_reason', '')
                ])
            
            table = Table(trade_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Son Trade'ler", getSampleStyleSheet()['Heading2']))
            story.append(table)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            self.logger.error(f"PDF trade bölümü ekleme hatası: {e}")
    
    def _add_monte_carlo_section(self, story, monte_carlo_results: Dict[str, Any]):
        """PDF Monte Carlo bölümü ekle"""
        try:
            statistical_analysis = monte_carlo_results.get('statistical_analysis', {})
            
            mc_data = [["Metrik", "Ortalama", "P5", "P95"]]
            
            for metric, stats in statistical_analysis.items():
                if isinstance(stats, dict) and 'mean' in stats:
                    mc_data.append([
                        metric,
                        f"{stats.get('mean', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p5', 0):.4f}",
                        f"{stats.get('percentiles', {}).get('p95', 0):.4f}"
                    ])
            
            table = Table(mc_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Monte Carlo Analizi", getSampleStyleSheet()['Heading2']))
            story.append(table)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            self.logger.error(f"PDF Monte Carlo bölümü ekleme hatası: {e}")

# Global rapor üretici
report_generator = ReportGenerator()




